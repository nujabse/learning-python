# -*- coding:utf-8 -*-

import urllib2
import sys
import re
import regex
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
import requests
from lxml import html
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup




class hustPhysicsTeachers:
    def __init__(self):
        reload(sys)
        sys.setdefaultencoding('utf-8')
        # print sys.stdin.encoding
        self.num = 900
        self.total = 0
        self.name = []
        self.email = []
        self.phone = []
        self.working_degree1 = []
        # self.content = []

    def getpage(self, num):
        try:
            user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
            headers = {'User-Agent': user_agent}
            url = 'http://phys.hust.edu.cn/xszdwjs/' + str(num) + '.htm'  # xi shi zi dui wu jie shao
            request = urllib2.Request(url, headers=headers)
            response = urllib2.urlopen(request)
            content = response.read().decode('utf-8')
            return content   # 返回content 交给其他函数调用

        except urllib2.URLError, e:
            if hasattr(e, "code"):
                print e.code
            if hasattr(e, "reason"):
                print u"错误原因", e.reason

    def getInformation(self):
        content = self.getpage(self.num)
        total = 0  # 老师总人数
        pattern_teacher = re.compile('<strong>(.*?)</strong>.*?')
        pattern_contact = re.compile('[a-z0-9._%+-]+@[A-Z0-9.-]+\\.[a-z]+\\.[a-z]+\\.*', re.IGNORECASE)
        pattern_phone = re.compile('\d{3}-\d{8}|\d{4}-\d{7}')
        pattern_working_degree = regex.compile(u'[教授]', re.UNICODE)
        if not content:
            self.num += 1
            print ("page %d not found" % self.num)
        else:
            teachers = re.findall(pattern_teacher, content)
            self.name.append(teachers[0])
            print teachers[0]
            contacts = re.findall(pattern_contact, content)
            phone = re.findall(pattern_phone, content)
            working_degree = regex.findall(pattern_working_degree, content)
            # 防止找不到联系而退出循环
            if not contacts:
                print u"没有联系方式!\n"
                self.email.append("None")
            else:
                print contacts[0]
                self.email.append(contacts[0])

            if not phone:
                print u"没有电话号码!\n"
                self.phone.append("None")
            else:
                print phone[0]
                self.phone.append(phone[0])
            # if not working_degree:
            #     print u"没有职称！\n"
            #     self.working_degree.append("None")
            # else:
            #     print working_degree[0]
            #     self.working_degree.append(working_degree[0])

    # def degreeAnalayze(self):
    #     soup = BeautifulSoup(self.getpage(self.num))
    #     # print(soup.get_text())
    #     degree = soup.find_all(text=u"职称")
    #     if degree:
    #         print degree[0]

    def get_professors(self):
        """
        获取教授介绍页信息，记得加self
        :return:
        """
        links = ['http://phys.hust.edu.cn/xszdwjs/index.htm', 'http://phys.hust.edu.cn/xszdwfjs/index.htm', 'http://phys.hust.edu.cn/xjiangshi/index.htm', 'http://phys.hust.edu.cn/xszdwjsry/index.htm']
        page = requests.get(links[2])
        tree = html.fromstring(page.text)
        for i in range(1, 70):

            index = '//*[@id="lbrc_1"]/ul/li[' + str(i) + ']/span/a/text()'
            # //*[@id="lbrc_1"]/table/tbody/tr[1]/td[1]/a
            intro_raw = tree.xpath(index)
            print intro_raw

            for i in intro_raw:
                intro = i.encode('utf-8')
            print intro
            self.working_degree1.append(intro)






    def writeToExcel(self):
        """
        存储采集数据到excel
        :return:
        """
        wb = load_workbook(u"数据采集.xlsx")
        ws1 = wb[u"导师"]
        cell_range = ws1['A1': 'D180']
        # ws1.title = u"导师"
        colA = ws1['A']
        colB = ws1['B']
        colC = ws1['C']
        colD = ws1['D']
        for i in range(len(self.name)):
            index = 'B' + str(i+2)
            ws1[index] = self.name[i]
        print("写入名称成功")
        for i in range(len(self.email)):
            index = 'C' + str(i+2)
            ws1[index] = self.email[i]
        print("写入邮箱成功")
        for i in range(len(self.phone)):
            index = 'D' + str(i+2)
            ws1[index] = self.phone[i]
        print("写入电话成功")
        for i in range(len(self.working_degree1)):
            index = 'G' + str(i+2)
            ws1[index] = self.working_degree1[i]
        print("写入讲师成功")



        # ws1['A1'] = u"编号"
        # ws1['B1'] = u"导师"
        # ws1['C1'] = u"邮箱"
        # ws1['D1'] = u"电话"
        # ws1['E1'] = u"职位"

        # for i in range(2, 200):
        #     print i
        #     index = 'A' + str(i)
        #     ws1[index] = i-1
        #     print ws1[index]
        wb.save(u"数据采集.xlsx")

    def statistics(self):
        """
        统计人数
        """
        if self.getpage(self.num):
            # 计算有多少个老师
            self.total += 1

    def test(self):
        # while self.num < 1054:
            # self.getpage(self.num)
            # self.statistics()
            # self.num += 1
            # print "*"*30
            # # if self.getpage(self.num):
            # #     self.degreeAnalayze()
            # print("checking page %d\n" % self.num)
            # self.getInformation()

        self.get_professors()
        # print("总共有 %d 个老师\n" % self.total)
        self.writeToExcel()
        # print self.name
        # print "*" * 30
        # print self.email
        # print "*" * 30
        # print self.phone

hustPhysicsTeachers().test()


# except urllib2.URLError, e:
#     if hasattr(e, "code"):
#         print e.code
#     if hasattr(e, "reason"):
#         print u"错误原因", e.reason
# print num
# print content

