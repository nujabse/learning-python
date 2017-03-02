# -*- coding:utf-8 -*-
import urllib2
import re
import sys
from lxml import etree
from openpyxl import load_workbook
from openpyxl.compat import range
import time
from socket import error as SocketError
import errno
import random

reload(sys)
sys.setdefaultencoding('utf-8')


class xiamiCounter():
    def __init__(self):
        self.headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
                    'Referer':'http://m.xiami.com/',
                    'Connection': 'keep-alive',

                   'Cookie':'unsign_token=4f23ab3b94ae95850c9c5da7aef287e2; bdshare_firstime=1460894873450; cna=aSmVD89xbyQCAcdTMxgCkCtM; gid=146233966652053; CNZZDATA921634=cnzz_eid%3D1056600907-1470359568-%26ntime%3D1472796701; CNZZDATA2629111=cnzz_eid%3D1838752247-1470359096-%26ntime%3D1472796766; _xiamitoken=d1fb6eac5ee905225c46226aef039c7b; login_method=emaillogin; member_auth=1zmQGYxLvj0xivSXS41me3AW5rCATDfXlokE27Aq5QFydYxYNYOrkKuTQApN3iSSq2FCRfnEhWQSRr0; user=22156413%22mathholic%22images%2Favatar_new%2F443%2F22156413_1394536403_1.jpg%222%2219655%22%3Ca+href%3D%27%2Fwebsitehelp%23help9_3%27+%3EDo%E2%80%A2%3C%2Fa%3E%22179%2267%2211887%2215e6c2b9dc%221488254727; _m_h5_tk=7c41b5eb2452b26db16b5b747b515b9c_1488283863638; _m_h5_tk_enc=15ff1a0b4a8465622b67a195472776fe; sec=58b57dd3547bf80f1f4a31a21817dcede3d13cc6; t_sign_auth=1; l=Ar29SsHoaUXW1EPeOTyxcTHnTRO3X/G5; isg=AoaGbUI5IiROGvZBVu5Xyu5Y13U5A8qhmkUoi3Cvd6mRcyaN2XcasWxDJRhF'
                   }
        self.workbook = load_workbook("data.xlsx")

    def getdata(self, lower_limit, upper_limit):
        """由于虾米部分音乐下架，可能部分音乐数据无法查证"""
        for i in range(lower_limit, upper_limit):
            try:
                url = 'http://www.xiami.com/count/getplaycount?id=' + str(i) + '&type=song&_xiamitoken=d1fb6eac5ee905225c46226aef039c7b'  # song ID
                # song_url = 'http://www.xiami.com/song/' + str(i)
                request = urllib2.Request(url, headers=self.headers)
                response = urllib2.urlopen(request)
                content = response.read().decode('utf-8')
                pattern_count = re.compile('[:](\d{1,100000000})[,]')
                count = re.findall(pattern_count, content)
                #
                # request_song = urllib2.Request(song_url, headers=headers)
                # response_song = urllib2.urlopen(request_song)
                # content_song = response_song.read().decode('utf-8')
                # page = etree.HTML(content_song)
                # index_song = '//*[@id="title"]/h1/text()'

                # name = page.xpath(index_song)

                # song = name
                if count:
                    # counter.append(count[0])
                    self.writeToExcel(i+1, 4, count[0])
                else:
                    self.writeToExcel(i+1, 4, 0)

                self.writeToExcel(i+1, 2, i)
                # ids.append(id)
                # title.append(name)
                print ("song number %d  play count %s" % (i, count))

            except SocketError as e:
                if e.errno != errno.ECONNRESET:
                    raise  # Not error we are looking for
                pass  # Handle error here.

            except urllib2.URLError, e:
                if hasattr(e, "code"):
                    print e.code
                if hasattr(e, "reason"):
                    print "reason", e.reason

            if i % 299 == 0:
                print("休息一下。。。。。")
                self.workbook.save("data.xlsx")
                print("保存数据成功！")
                time.sleep(random.randrange(0,5))  # 设置时间间隔为299秒


    def writeToExcel(self, rows, col, data):
        """
        存储采集数据到excel
        :return:
        """
        ws1 = self.workbook["xiami"]
        ws1['A1'] = u"编号"
        ws1['B1'] = "ids"
        ws1['C1'] = "title"
        ws1['D1'] = "counter"
        ws1['E1'] = "artist"
        # colA = ws1['A']
        # colB = ws1['B']
        # colC = ws1['C']
        # colD = ws1['D']
        ws1.cell(row=rows, column=col, value=int(data))
        # for i in range(len(ids)):
        #     index = 'B' + str(i + 2)
        #     ws1[index] = ids[i]
        # print(u"写入id.....")
        # # for i in range(len(title)):
        # #     index = 'C' + str(i + 2)
        # #     ws1[index] = title[i]
        # # print(u"写入title成功")
        # for i in range(len(counter)):
        #     index = 'D' + str(i + 2)
        #     ws1[index] = counter[i]

int1 = int(raw_input("please input lower limit: "))
int2 = int(raw_input("please input upper limit: "))
xiamiCounter().getdata(int1, int2)
print(u"采集数据完成！")




