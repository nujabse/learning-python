# -*- coding:utf-8 -*-
import re
import sys
import requests
from lxml import etree
from lxml import html
from openpyxl import load_workbook
from openpyxl.compat import range
import time
from socket import error as SocketError
import errno
import random

reload(sys)
sys.setdefaultencoding('utf-8')


class album():
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
            'Referer': 'http://m.xiami.com/',
            'Connection': 'keep-alive'

            # 'Cookie': 'unsign_token=4f23ab3b94ae95850c9c5da7aef287e2; bdshare_firstime=1460894873450; cna=aSmVD89xbyQCAcdTMxgCkCtM; gid=146233966652053; CNZZDATA921634=cnzz_eid%3D1056600907-1470359568-%26ntime%3D1472796701; CNZZDATA2629111=cnzz_eid%3D1838752247-1470359096-%26ntime%3D1472796766; _xiamitoken=d1fb6eac5ee905225c46226aef039c7b; login_method=emaillogin; member_auth=1zmQGYxLvj0xivSXS41me3AW5rCATDfXlokE27Aq5QFydYxYNYOrkKuTQApN3iSSq2FCRfnEhWQSRr0; user=22156413%22mathholic%22images%2Favatar_new%2F443%2F22156413_1394536403_1.jpg%222%2219655%22%3Ca+href%3D%27%2Fwebsitehelp%23help9_3%27+%3EDo%E2%80%A2%3C%2Fa%3E%22179%2267%2211887%2215e6c2b9dc%221488254727; _m_h5_tk=7c41b5eb2452b26db16b5b747b515b9c_1488283863638; _m_h5_tk_enc=15ff1a0b4a8465622b67a195472776fe; sec=58b57dd3547bf80f1f4a31a21817dcede3d13cc6; t_sign_auth=1; l=Ar29SsHoaUXW1EPeOTyxcTHnTRO3X/G5; isg=AoaGbUI5IiROGvZBVu5Xyu5Y13U5A8qhmkUoi3Cvd6mRcyaN2XcasWxDJRhF'
            }
        self.workbook = load_workbook("album.xlsx")

    def getdata(self, lower_limit, upper_limit):
        for num in range(lower_limit, upper_limit):
            try:
                url = 'http://www.xiami.com/album/' + str(num)
                page = requests.get(url, headers=self.headers, timeout=5)
                tree = html.fromstring(page.text)
                index1 = '//*[@id="sidebar"]/div[1]/ul/li[1]/text()'
                index2 = '//*[@id="sidebar"]/div[1]/ul/li[3]/a/i/text()'
                index3 = '//*[@id="album_rank"]/p/em/text()'
                index5 = '//*[@id="title"]/h1/text()'
                # index6 = '//*[@id="album_info"]/table/tbody/tr[1]/td[2]/a/text()'
                play = tree.xpath(index1)
                comment = tree.xpath(index2)
                rating = tree.xpath(index3)
                title = tree.xpath(index5)
                # artist = tree.xpath(index6)
                if play:
                    if rating:
                        print "Album ID %d play count: %s comment: %s rating: %s title: %s " % (num, play[0], comment[0], rating[0], title[0])
                        self.writeToExcel(num, 1, num)
                        self.writeToExcel(num, 2, play[0])
                        self.writeToExcel(num, 3, rating[0])
                        self.writeToExcel(num, 4, comment[0])
                        self.writeToExcel(num, 5, title[0])
                    else:
                        print(u"没有评价！")
                        self.writeToExcel(num, 1, num)
                        self.writeToExcel(num, 2, play[0])
                        self.writeToExcel(num, 3, 0)
                        self.writeToExcel(num, 4, comment[0])
                        self.writeToExcel(num, 5, title[0])
                else:
                    print "Page number %d not exist!!" % num
                    pass
                print(u"采集数据中。。。。。。。。。。。")

                if num % 199 == 0 or num == upper_limit - 1:
                    self.workbook.save("album.xlsx")
                    print("保存数据成功！")

            except SocketError as e:
                if e.errno != errno.ECONNRESET:
                    raise  # Not error we are looking for
                pass  # Handle error here.

            except requests.ConnectionError as e:
                if e.errno == 104:
                    num -= 1
                    continue
            except requests.ReadTimeout:
                print "请求频繁， 小睡一下再重试。。。。。"
                time.sleep(10)
                num -= 1
                continue

            if num % 77 == 1:
                print("休息一下。。。。。")
                time.sleep(random.randrange(0, 20))  # 设置时间间隔为299秒

            # if hasattr(e, "reason"):
            #     print "reason", e.reason



    def writeToExcel(self, rows, col, data):
        """
        存储采集数据到excel
        :return:
        """
        ws1 = self.workbook["album"]
        ws1['A1'] = "id"
        ws1['B1'] = "play"
        ws1['C1'] = "rating"
        ws1['D1'] = "comment"
        ws1['E1'] = "title"
        ws1['F1'] = "artist"
        if col == 3:        # add condition for string data
            ws1.cell(row=rows + 1, column=col, value=float(data))
        else:
            if col == 5:
                ws1.cell(row=rows + 1, column=col, value=data)
            else:
                ws1.cell(row=rows + 1, column=col, value=int(data))

int1 = int(raw_input("please input lower limit: "))
int2 = int(raw_input("please input upper limit: "))
album().getdata(int1, int2)
print(u"采集数据完成！")