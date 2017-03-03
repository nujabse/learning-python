# -*- coding:utf-8 -*-
import urllib2
import re
import sys
from lxml import etree
from openpyxl import load_workbook
from openpyxl.compat import range
import time
from socket import error as SocketError
import errno
import random

reload(sys)
sys.setdefaultencoding('utf-8')


class Bilicounter:
    def __init__(self):
        self.headers = {"Host": "api.bilibili.com",
                        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:51.0) Gecko/20100101 Firefox/51.0',
                        'Accept-Language': 'en-US,en;q=0.5',
                        'Referer': 'http://www.bilibili.com/video/av8898537/',
                        'Cookie': 'fts=1460870702; sid=bbmgsz04; rpdid=olmiqlmqqsdopmoswiiiw; buvid3=F78F18B2-38A8-4A81-A02B-9DB098A08A9724942infoc',
                        'Connection': 'keep-alive'}
        self.workbook = load_workbook("huster.xlsx")


    def getdata(self, lower_limit, upper_limit):
        for num in range(lower_limit, upper_limit):
            try:
                url = 'http://api.bilibili.com/archive_stat/stat?aid=' + str(num)
                request = urllib2.Request(url, headers=self.headers)
                response = urllib2.urlopen(request, timeout=6)
                content = response.read().decode('utf-8')
                pattern_view = re.compile("\"view\":(.*?),")
                pattern_danmaku = re.compile("\"danmaku\":(.*?),")
                pattern_reply = re.compile("\"reply\":(.*?),")
                pattern_favorite= re.compile("\"favorite\":(.*?),")
                pattern_coin = re.compile("\"coin\":(.*?),")
                pattern_share = re.compile("\"share\":(.*?),")
                pattern_his_rank = re.compile("\"his_rank\":(.*?)}")
                pattern_now_rank = re.compile("\"now_rank\":(.*?),")
                view = re.findall(pattern_view, content)
                danmaku = re.findall(pattern_danmaku, content)
                reply = re.findall(pattern_reply, content)
                favorite = re.findall(pattern_favorite, content)
                coin = re.findall(pattern_coin, content)
                share = re.findall(pattern_share, content)
                his_rank = re.findall(pattern_his_rank, content)
                now_rank = re.findall(pattern_now_rank, content)
                if view:
                    self.writeToExcel(num, 1, num)
                    self.writeToExcel(num, 2, view[0])
                    self.writeToExcel(num, 3, danmaku[0])
                    self.writeToExcel(num, 4, reply[0])
                    self.writeToExcel(num, 5, favorite[0])
                    self.writeToExcel(num, 6, coin[0])
                    self.writeToExcel(num, 7, share[0])
                    self.writeToExcel(num, 8, his_rank[0])
                    self.writeToExcel(num, 9, now_rank[0])
                    print(u"采集数据中。。。。。。。。。。。")

                    print("video id %d view count %s danmaku %s reply %s favorite %s coin %s share %s his_rank %s now_rank %s" % (num, view[0], danmaku[0], reply[0], favorite[0], coin[0], share[0], his_rank[0], now_rank[0]))
                else:
                    self.writeToExcel(num, 2, 0)

            except SocketError as e:
                if e.errno != errno.ECONNRESET:
                    raise  # Not error we are looking for
                pass  # Handle error here.

            except urllib2.URLError, e:
                if hasattr(e, "code"):
                    print e.code
                if hasattr(e, "reason"):
                    print "reason", e.reason
            if num % 299 == 0 or num == upper_limit - 1:
                print("休息一下。。。。。")
                self.workbook.save("huster.xlsx")   # 设置99秒保存一次
                print("存入数据成功")
                time.sleep(random.randrange(0, 15))  # 设置时间间隔为299秒



    def writeToExcel(self, rows, col, data):
        """
        存储采集数据到excel
        :return:
        """
        # wb = load_workbook("huster.xlsx")
        ws1 = self.workbook["bilibili"]
        # ws1['A1'] = "ID"
        # ws1['B1'] = "view"
        # ws1['C1'] = "danmaku"
        # ws1['D1'] = "reply"
        # ws1['E1'] = "favorite"
        # ws1['F1'] = "coin"
        # ws1['G1'] = "share"
        # ws1['H1'] = "History Rank"
        # ws1['I1'] = "Now Rank"
        if data == '"--"':
            ws1.cell(row=rows+1, column=col, value='Unknown')
        else:
            ws1.cell(row=rows+1, column=col, value=int(data))
        # for i in range(len(ids)):
        #     index = 'B' + str(i + 2)
        #     ws1[index] = ids[i]
        # print(u"写入id.....")
        # # for i in range(len(title)):
        # #     index = 'C' + str(i + 2)
        # #     ws1[index] = title[i]
        # # print(u"写入title成功")
        # for i in range(len(counter)):
        #     index = 'D' + str(i + 2)
        #     ws1[index] = counter[i]
        # wb.save("huster.xlsx")


        # for i in range(len(self.working_degree1)):
        #     index = 'G' + str(i + 2)
        #     ws1[index] = self.working_degree1[i]
        # print(u"写入讲师成功")

int1 = int(raw_input("please input lower limit: "))
int2 = int(raw_input("please input upper limit: "))
Bilicounter().getdata(int1, int2)
print(u"采集数据完成！")







