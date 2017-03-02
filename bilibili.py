# -*- coding:utf-8 -*-
import urllib2
import re
import sys
from lxml import etree
from openpyxl import load_workbook
from openpyxl.compat import range
import time
from socket import error as SocketError
import errno
import random

reload(sys)
sys.setdefaultencoding('utf-8')

headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
            'Referer':'http://www.bilibili.com/video/av8846918/',
            'Connection': 'keep-alive',

           'Cookie':'sid=lm09yj1b; fts=1460632018; time_tracker=20160616; buvid3=2AE5260B-E7C7-4D75-9BE1-62684F336E31956infoc; DedeUserID=29311993; DedeUserID__ckMd5=587efca938f0b74a; SESSDATA=cc6b3143%2C1490680144%2C933e5b5f; bili_jct=cbe1655d88fc1eda137eb6b02a3de006; _cnt_dyn=null; uTZ=-480; _cnt_pm=0; _cnt_notify=0; LIVE_LOGIN_DATA=eecd2fdfdf874d394b433f98c7bd822416c928ec; LIVE_LOGIN_DATA__ckMd5=d036889b161b1339; LIVE_BUVID=f947fa07510878c9abd57824c38fa3b6; LIVE_BUVID__ckMd5=9c951293b10e641c; user_face=http%3A%2F%2Fstatic.hdslb.com%2Fimages%2Fmember%2Fnoface.gif; rpdid=olmiqllmlqdopqqkkmkiw; purl_token=bilibili_1488337431; _dfcaptcha=854ce770b448867142655f4a639b9cae'
           }
def getdata(lower_limit, upper_limit):
    for num in range(lower_limit, upper_limit):
        try:
            url = 'http://api.bilibili.com/archive_stat/stat?aid=' + str(num)
            request = urllib2.Request(url, headers=headers)
            response = urllib2.urlopen(request)
            content = response.read().decode('utf-8')
            pattern_view = re.compile("\"view\":(.*?),")
            pattern_danmaku = re.compile("\"danmaku\":(.*?),")
            pattern_reply = re.compile("\"reply\":(.*?),")
            pattern_favorite= re.compile("\"favorite\":(.*?),")
            pattern_coin = re.compile("\"coin\":(.*?),")
            pattern_share = re.compile("\"share\":(.*?),")
            pattern_his_rank = re.compile("\"his_rank\":(.*?)}")
            pattern_now_rank = re.compile("\"now_rank\":(.*?),")
            view = re.findall(pattern_view, content)
            danmaku = re.findall(pattern_danmaku, content)
            reply = re.findall(pattern_reply, content)
            favorite = re.findall(pattern_favorite, content)
            coin = re.findall(pattern_coin, content)
            share = re.findall(pattern_share, content)
            his_rank = re.findall(pattern_his_rank, content)
            now_rank = re.findall(pattern_now_rank, content)
            if view:
                writeToExcel(num, 1, num)
                writeToExcel(num, 2, view[0])
                writeToExcel(num, 3, danmaku[0])
                writeToExcel(num, 4, reply[0])
                writeToExcel(num, 5, favorite[0])
                writeToExcel(num, 6, coin[0])
                writeToExcel(num, 7, share[0])
                writeToExcel(num, 8, his_rank[0])
                writeToExcel(num, 9, now_rank[0])
                print(u"采集数据中。。。。。。。。。。。")
                print("video id %d view count %s danmaku %s reply %s favorite %s coin %s share %s his_rank %s now_rank %s" % (num, view[0], danmaku[0], reply[0], favorite[0], coin[0], share[0], his_rank[0], now_rank[0]))
            else:
                view = 0

        except SocketError as e:
            if e.errno != errno.ECONNRESET:
                raise  # Not error we are looking for
            pass  # Handle error here.

        except urllib2.URLError, e:
            if hasattr(e, "code"):
                print e.code
            if hasattr(e, "reason"):
                print "reason", e.reason
        if num % 99 == 0:
            time.sleep(random.randrange(0,10))  # 设置时间间隔为19秒
            print("休息一下。。。。。")

def writeToExcel(rows, col, data):
    """
    存储采集数据到excel
    :return:
    """
    wb = load_workbook("huster.xlsx")
    ws1 = wb["bilibili"]
    # ws1['A1'] = "ID"
    # ws1['B1'] = "view"
    # ws1['C1'] = "danmaku"
    # ws1['D1'] = "reply"
    # ws1['E1'] = "favorite"
    # ws1['F1'] = "coin"
    # ws1['G1'] = "share"
    # ws1['H1'] = "History Rank"
    # ws1['I1'] = "Now Rank"
    if data == '"--"':
        ws1.cell(row=rows+1, column=col, value='Unknown')
    else:
        ws1.cell(row=rows+1, column=col, value=int(data))
    # for i in range(len(ids)):
    #     index = 'B' + str(i + 2)
    #     ws1[index] = ids[i]
    # print(u"写入id.....")
    # # for i in range(len(title)):
    # #     index = 'C' + str(i + 2)
    # #     ws1[index] = title[i]
    # # print(u"写入title成功")
    # for i in range(len(counter)):
    #     index = 'D' + str(i + 2)
    #     ws1[index] = counter[i]
    wb.save("huster.xlsx")


    # for i in range(len(self.working_degree1)):
    #     index = 'G' + str(i + 2)
    #     ws1[index] = self.working_degree1[i]
    # print(u"写入讲师成功")

int1 = int(raw_input("please input lower limit: "))
int2 = int(raw_input("please input upper limit: "))
getdata(int1, int2)








