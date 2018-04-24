import vobject
from openpyxl import load_workbook

wb = load_workbook(filename = 'contacts.xlsx')
ws = wb['Sheet1']
name_column = ws['C3': 'C39']
phone_column = ws['K3': 'K39']
adr_column = ws['M3': 'M39']
org_column = ws['I3': 'I39']
note_column = ws['L3': 'L39']
for cell in adr_column:
    print(cell[0].value)


# j.add('n')
# j.n.value = vobject.vcard.Name( family='Harris', given='Jeffrey' )

# j.add('fn').value = "Jeffery Harris"
# j.prettyPrint()
with open('output.vcf', 'w') as output:
    for i in range(37):
        j = vobject.vCard()
        j.add('fn').value = name_column[i][0].value
        j.add('tel').value = str(phone_column[i][0].value)
        j.add('label').value = '17硕三班'
        org = org_column[i][0].value + note_column[i][0].value
        j.add('org').value = org
        # Address name must be written in this way, or it will warn you missing box attribute
        #  when do a serialize()
        j.add('adr').value = vobject.vcard.Address((adr_column[i][0].value))
        j.adr.type_param = 'HOME'
        # j.add('note').value =note_column[i][0].value

        j.serialize()
        i = i + 1
        output.write(j.serialize())    # seems like this module can only write one contact at a time

